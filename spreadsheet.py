import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from links import build_search_url


HEADERS = ["Address", "Date of Sale", "Price (€)", "Property Type", "Search", "Comments"]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def generate_spreadsheet(matches: pd.DataFrame, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PPR Results"

    # Write header
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Write rows
    for row_idx, (_, row) in enumerate(matches.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=str(row.get("Address", "")).strip())
        ws.cell(row=row_idx, column=2, value=str(row.get("Date of Sale (dd/mm/yyyy)", "")).strip())

        price = row.get("Price (€)", "")
        try:
            price_val = float(price)
            cell = ws.cell(row=row_idx, column=3, value=price_val)
            cell.number_format = '#,##0.00'
        except (ValueError, TypeError):
            ws.cell(row=row_idx, column=3, value=str(price).strip())

        ws.cell(row=row_idx, column=4, value=str(row.get("Description of Property", "")).strip())

        address = str(row.get("Address", "")).strip()
        search_cell = ws.cell(row=row_idx, column=5, value="Search")
        search_cell.hyperlink = build_search_url(address)
        search_cell.font = Font(color="0563C1", underline="single")

        ws.cell(row=row_idx, column=6, value="")  # Comments — left blank for user

    # Auto-fit column widths
    col_min_widths = [50, 14, 16, 30, 10, 30]
    for col_idx, min_width in enumerate(col_min_widths, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, 80)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
