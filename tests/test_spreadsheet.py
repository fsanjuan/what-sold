import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook

from spreadsheet import generate_spreadsheet, HEADERS


@pytest.fixture
def sample_matches():
    return pd.DataFrame([
        {
            "Address": "APT 10, BLOCK A, SMITHFIELD MARKET, DUBLIN 7",
            "Date of Sale (dd/mm/yyyy)": "15/06/2024",
            "Price (€)": 350000.0,
            "Description of Property": "Second-Hand Dwelling house /Apartment",
        },
        {
            "Address": "22 SMITHFIELD VILLAGE, DUBLIN 7",
            "Date of Sale (dd/mm/yyyy)": "01/03/2024",
            "Price (€)": 420000.0,
            "Description of Property": "Second-Hand Dwelling house /Apartment",
        },
    ])


@pytest.fixture
def output_xlsx(sample_matches):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    generate_spreadsheet(sample_matches, path)
    yield path
    os.unlink(path)


class TestSpreadsheetStructure:
    def test_file_is_created(self, output_xlsx):
        assert os.path.exists(output_xlsx)

    def test_correct_sheet_name(self, output_xlsx):
        wb = load_workbook(output_xlsx)
        assert "PPR Results" in wb.sheetnames

    def test_header_row_matches(self, output_xlsx):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        actual_headers = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
        assert actual_headers == HEADERS

    def test_row_count(self, output_xlsx, sample_matches):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        # 1 header + N data rows
        assert ws.max_row == 1 + len(sample_matches)

    def test_freeze_panes(self, output_xlsx):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        assert ws.freeze_panes == "A2"


class TestSpreadsheetContent:
    def test_address_written(self, output_xlsx, sample_matches):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        assert cell_value == sample_matches.iloc[0]["Address"]

    def test_date_written(self, output_xlsx, sample_matches):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "15/06/2024"

    def test_price_written_as_number(self, output_xlsx, sample_matches):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == 350000.0

    def test_comments_column_is_empty(self, output_xlsx, sample_matches):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            assert ws.cell(row=row, column=5).value in (None, "")

    def test_header_is_bold(self, output_xlsx):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold

    def test_invalid_price_written_as_string(self, tmp_path):
        df = pd.DataFrame([{
            "Address": "1 Test St",
            "Date of Sale (dd/mm/yyyy)": "01/01/2024",
            "Price (€)": "N/A",
            "Description of Property": "Test",
        }])
        path = str(tmp_path / "out.xlsx")
        generate_spreadsheet(df, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "N/A"

    def test_empty_dataframe(self, tmp_path):
        df = pd.DataFrame(columns=["Address", "Date of Sale (dd/mm/yyyy)", "Price (€)", "Description of Property"])
        path = str(tmp_path / "empty.xlsx")
        generate_spreadsheet(df, path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.max_row == 1  # header only


class TestSpreadsheetFormatting:
    def test_column_widths_set(self, output_xlsx):
        wb = load_workbook(output_xlsx)
        ws = wb.active
        for col in range(1, len(HEADERS) + 1):
            letter = ws.cell(row=1, column=col).column_letter
            assert ws.column_dimensions[letter].width > 0
